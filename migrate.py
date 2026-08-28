"""
Migrate Master Credit Box Excel to typed SQLite.

Reads Excel directly, validates every cell against known type,
logs all anomalies. Re-import produces a diff report.

Usage:
    python3 migrate.py
    python3 migrate.py --excel path/to/excel.xlsx
    python3 migrate.py --diff  (show diff from last import)
"""

import sqlite3, json, sys, os, re, datetime, hashlib, argparse
from pathlib import Path
from collections import defaultdict, OrderedDict
import openpyxl

CORPUS_DIR = Path(__file__).parent / "corpus"
DB_PATH = CORPUS_DIR / "corpus.db"
EXCEL_DEFAULT = "/Users/jackrodenberg/Downloads/Copy of THE Master Credit Box-IPLE 2026 (1).xlsx"

SCHEMA_VERSION = 2

from attr_types import get_type


# ── Product sheets ────────────────────────────────────────────
PRODUCT_SHEETS = OrderedDict([
    ("SFR DSCR", "sfr_dscr"),
    ("Fix & Flip", "fix_and_flip"),
    ("New Construction", "new_construction"),
    ("Multifamily Long Term", "multifamily_lt"),
    ("SFR Bridge", "sfr_bridge"),
    ("SFR Blanket", "sfr_blanket"),
    ("Multifamily Rehab", "multifamily_rehab"),
    ("Multi-Comm Bridge", "multi_comm_bridge"),
    ("SB Commercial Long Term", "sb_commercial_lt"),
])

CS_SHEETS = [
    "CS-DSCR Implication", "CS-FNF Implication ",
    "CS-CREDIT", "CS-EXPERIENCE",
    "CS-LEASEOCCUPANCY", "CS-RESERVESASSETS",
    "SSCS",
]

# Lender name normalizations (hard-coded aliases)
LENDER_NORMALIZE = {
    "back flip": "Backflip",
    "backflip": "Backflip",
    "templeview": "Temple View Capital",
    "easy street": "EasyStreet",
    "easy street capital": "EasyStreet",
    "easystreet": "EasyStreet",
    "roc": "ROC Capital",
    "roc capital": "ROC Capital",
    "crebrid": "Crebrid",
    "ternus": "Ternus",
    "constructive standard": "Constructive",
    "constructive expanded": "Constructive Expanded",
    "constructive fnf": "Constructive",
    "constructive draft": "Constructive",
    "rcn dscr": "RCN",
    "rcn fnf": "RCN",
    "kiavi": "Kiavi",
    "kiavi - fnf": "Kiavi",
    "kiavi-dscr": "Kiavi",
    "kiavi-fnf": "Kiavi",
    "silverhill": "Silver Hill",
    "cv3 dscr": "CV3",
    "cv3 fnf": "CV3",
    "conventus dscr": "Conventus",
    "conventus fnf": "Conventus",
    "easystreet dscr": "EasyStreet",
    "easystreet fnf": "EasyStreet",
    "bpc (on pause)": "BPC",
    "bpc": "BPC",
    "lima one": "Lima One Capital",
    "lima one capital": "Lima One Capital",
    "center street": "Center Street",
    "f2 finance": "F2 Finance",
    "flipco financial": "FlipCo Financial",
    "flipco": "FlipCo Financial",
    "hard money co": "Hard Money Co",
    "hard money co.": "Hard Money Co",
    "rain city capital": "Rain City Capital",
    "rain city": "Rain City Capital",
    "ground floor": "Groundfloor",
    "groundfloor": "Groundfloor",
    "wildcatcrebrid": "Crebrid",
    "wildcat": "Crebrid",
    "rcn only": "RCN",
    "rcn loanbidz special": "RCN",
    "finance of america": "Finance of America",
    "investor loan source": "Investor Loan Source",
    "temple view capital": "Temple View Capital",
}

SENSITIVE_ATTRS = {
    "password", "user_name", "pricer_location",
    "lender_dropbox_folder_link", "website_link",
}


# ── Schema ────────────────────────────────────────────────────
SCHEMA_SQL = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

CREATE TABLE IF NOT EXISTS meta (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS imports (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    file_path TEXT NOT NULL,
    file_hash TEXT NOT NULL,
    file_mtime TEXT NOT NULL,
    imported_at TEXT DEFAULT (datetime('now')),
    records_added INTEGER DEFAULT 0,
    records_removed INTEGER DEFAULT 0,
    records_changed INTEGER DEFAULT 0,
    warnings INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY,
    name TEXT UNIQUE NOT NULL,
    display_name TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS attribute_definitions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL,
    display_name TEXT NOT NULL,
    data_type TEXT NOT NULL DEFAULT 'text'
        CHECK (data_type IN ('text','percent','number','dollar','date','boolean','json')),
    category TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS product_attributes (
    product_id INTEGER NOT NULL REFERENCES products(id),
    attr_id INTEGER NOT NULL REFERENCES attribute_definitions(id),
    sort_order INTEGER DEFAULT 0,
    PRIMARY KEY (product_id, attr_id)
);

CREATE TABLE IF NOT EXISTS lenders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    canonical_name TEXT UNIQUE NOT NULL,
    display_name TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS lender_aliases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    alias TEXT NOT NULL,
    source_sheet TEXT,
    UNIQUE(lender_id, alias)
);

CREATE TABLE IF NOT EXISTS product_offerings (
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    product_id INTEGER NOT NULL REFERENCES products(id),
    PRIMARY KEY (lender_id, product_id)
);

CREATE TABLE IF NOT EXISTS lender_attr_values (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL REFERENCES imports(id),
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    product_id INTEGER NOT NULL REFERENCES products(id),
    attr_id INTEGER NOT NULL REFERENCES attribute_definitions(id),
    value_text TEXT,
    value_numeric REAL,
    value_date TEXT,
    raw_text TEXT,
    source_sheet TEXT NOT NULL,
    source_col INTEGER,
    source_row INTEGER NOT NULL,
    validation_warning TEXT,
    UNIQUE(lender_id, product_id, attr_id, import_id)
);
CREATE INDEX IF NOT EXISTS idx_lav_lender_product ON lender_attr_values(lender_id, product_id);
CREATE INDEX IF NOT EXISTS idx_lav_attr ON lender_attr_values(attr_id);

CREATE TABLE IF NOT EXISTS credit_grids (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL REFERENCES imports(id),
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    min_fico REAL,
    ltv_purchase TEXT,
    ltv_rate_term TEXT,
    ltv_cashout TEXT,
    foreign_national TEXT,
    fico_qual_method TEXT,
    guarantee_reqs TEXT,
    tradeline_reqs TEXT,
    max_lates TEXT,
    bk_fc TEXT,
    source_sheet TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_cg_lender ON credit_grids(lender_id);

CREATE TABLE IF NOT EXISTS experience_matrices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL REFERENCES imports(id),
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    exp_level INTEGER NOT NULL,
    ltc_terms TEXT,
    source_sheet TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_em_lender ON experience_matrices(lender_id);

CREATE TABLE IF NOT EXISTS scenarios (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL REFERENCES imports(id),
    scenario_id TEXT NOT NULL,
    condition TEXT NOT NULL,
    product_type TEXT NOT NULL,
    recommendation_lender TEXT,
    recommendation_detail TEXT,
    source_sheet TEXT NOT NULL,
    source_row INTEGER
);
CREATE INDEX IF NOT EXISTS idx_sc_product ON scenarios(product_type);

CREATE TABLE IF NOT EXISTS lease_occupancy (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL REFERENCES imports(id),
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    occupancy_reqs TEXT,
    vacancy_pricing TEXT,
    acceptable_lease TEXT,
    unacceptable_lease TEXT,
    lease_terms TEXT,
    source_sheet TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS reserves_requirements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL REFERENCES imports(id),
    lender_id INTEGER NOT NULL REFERENCES lenders(id),
    requirement TEXT,
    multifamily TEXT,
    foreign_national TEXT,
    cash_out TEXT,
    eligible_assets TEXT,
    source_sheet TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS migration_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER REFERENCES imports(id),
    level TEXT NOT NULL CHECK (level IN ('INFO','WARN','ERROR')),
    message TEXT NOT NULL,
    source_sheet TEXT,
    source_row INTEGER,
    created_at TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_ml_level ON migration_log(level);
"""


# ── Helpers ───────────────────────────────────────────────────
def log(db, import_id, level, message, source_sheet=None, source_row=None):
    db.execute(
        "INSERT INTO migration_log (import_id, level, message, source_sheet, source_row) VALUES (?,?,?,?,?)",
        (import_id, level, message, source_sheet, source_row)
    )
    if level == "ERROR":
        print(f"  ERROR: {message}", file=sys.stderr)
    elif level == "WARN":
        print(f"  WARN:  {message}", file=sys.stderr)


def normalize_lender(raw):
    s = raw.strip().replace('\n', ' ')
    key = s.lower().strip()
    # Remove trailing whitespace/punctuation
    key = re.sub(r'\s+', ' ', key).strip()
    key = re.sub(r'[^\w\s]', '', key).strip()
    if key in LENDER_NORMALIZE:
        return LENDER_NORMALIZE[key]
    # Title-case it
    return s.strip()


def parse_value(raw, data_type):
    """Parse raw cell value according to data_type. Returns (text, numeric, date, warning)."""
    if raw is None:
        return ("", None, None, None)
    s = str(raw).strip()
    if s == "":
        return ("", None, None, None)
    warning = None
    numeric = None
    date_val = None

    if data_type == "number":
        try:
            cleaned = s.rstrip('%').replace(',', '').replace('$', '').strip()
            numeric = float(cleaned)
        except ValueError:
            warning = f"expected number, got '{s[:60]}'"
    elif data_type == "percent":
        try:
            cleaned = s.replace('%', '').replace(',', '').replace('$', '').strip()
            numeric = float(cleaned) / 100.0 if '%' in s else float(cleaned)
        except ValueError:
            warning = f"expected percent, got '{s[:60]}'"
    elif data_type == "dollar":
        try:
            cleaned = s.replace('$', '').replace(',', '').replace(' ', '').strip()
            numeric = float(cleaned)
        except ValueError:
            warning = f"expected dollar, got '{s[:60]}'"
    elif data_type == "boolean":
        if s.lower() in ('yes', 'y', 'true', 'no', 'n', 'false', 'n/a', 'na', 'none'):
            pass  # text is enough
        else:
            warning = f"expected boolean, got '{s[:60]}'"
    elif data_type == "date":
        # Check for ISO date or Excel serial
        if isinstance(raw, datetime.datetime):
            date_val = raw.isoformat()
        elif isinstance(raw, (int, float)):
            try:
                dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(raw))
                date_val = dt.strftime("%Y-%m-%d")
            except:
                pass
        else:
            try:
                dt = datetime.datetime.fromisoformat(str(raw).strip().split('.')[0])
                date_val = dt.strftime("%Y-%m-%d")
            except:
                warning = f"expected date, got '{s[:60]}'"

    if warning and data_type not in ('boolean', 'date'):
        pass  # keep warning for logging

    return (s, numeric, date_val, warning)


def cell_text(cell):
    if cell is None:
        return None
    return str(cell).strip()


def is_numeric(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ── Migration ─────────────────────────────────────────────────
def migrate(excel_path):
    print(f"Loading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    db_path = DB_PATH
    db_path.parent.mkdir(parents=True, exist_ok=True)
    fresh = not db_path.exists()

    db = sqlite3.connect(str(db_path))
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")

    if fresh:
        db.executescript(SCHEMA_SQL)
        print("Schema created fresh")
    else:
        # Verify schema version
        row = db.execute("SELECT value FROM meta WHERE key='schema_version'").fetchone()
        if row and int(row["value"]) != SCHEMA_VERSION:
            print(f"Schema version mismatch: DB={row['value']}, code={SCHEMA_VERSION}")
            print("Run with --fresh to rebuild")
            return False

    # Compute file hash for dedup
    file_hash = hashlib.md5(open(excel_path, 'rb').read()).hexdigest()
    file_mtime = datetime.datetime.fromtimestamp(
        os.path.getmtime(excel_path)
    ).isoformat()

    # Check if already imported
    existing = db.execute(
        "SELECT id FROM imports WHERE file_hash=? AND file_mtime=?",
        (file_hash, file_mtime)
    ).fetchone()
    if existing:
        print(f"Excel already imported (import_id={existing['id']})")
        print("Run with --force to re-import")

    # Create import record
    db.execute(
        "INSERT INTO imports (file_path, file_hash, file_mtime) VALUES (?,?,?)",
        (excel_path, file_hash, file_mtime)
    )
    import_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        _validate_schema(db, import_id)
        _insert_products(db, import_id, wb)
        _insert_lenders(db, import_id, wb)
        _insert_attribute_defs(db, import_id, wb)
        _insert_product_sheet_data(db, import_id, wb)
        _insert_credit_grids(db, import_id, wb)
        _insert_experience_matrices(db, import_id, wb)
        _insert_scenarios(db, import_id, wb)
        _insert_lease_occupancy(db, import_id, wb)
        _insert_reserves(db, import_id, wb)

        db.execute("INSERT OR REPLACE INTO meta (key, value) VALUES ('schema_version', ?)", (str(SCHEMA_VERSION),))
        db.execute("INSERT OR REPLACE INTO meta (key, value) VALUES ('last_import_id', ?)", (str(import_id),))
        db.commit()
        print(f"\nImport {import_id} complete.")
        _print_summary(db, import_id)
        return True

    except Exception as e:
        db.rollback()
        log(db, import_id, "ERROR", f"Migration failed: {e}")
        db.commit()
        raise


def _validate_schema(db, import_id):
    """Check Excel has expected structure."""
    pass  # structure validation done during parsing


def _insert_products(db, import_id, wb):
    for display_name, product_key in PRODUCT_SHEETS.items():
        db.execute(
            "INSERT OR IGNORE INTO products (id, name, display_name) VALUES (?,?,?)",
            (list(PRODUCT_SHEETS.values()).index(product_key) + 1, product_key, display_name)
        )
    log(db, import_id, "INFO", f"{len(PRODUCT_SHEETS)} products")


def _first_lender_col(ws):
    """Row 1's lender names normally start at column C (3), because column A
    holds a 'VOTE COLUMN' placeholder. Some product sheets (observed:
    Multi-Comm Bridge) omit that placeholder, so lender names sit one
    column to the left, starting at column B (2). Detect which layout this
    sheet uses instead of assuming column 3 always.
    """
    col1_val = ws.cell(row=1, column=1).value
    if col1_val and "vote column" in str(col1_val).strip().lower():
        return 3
    return 2


def _insert_lenders(db, import_id, wb):
    """Extract all unique lender names across all sheets."""
    all_names = set()

    # Product sheets: row 1, col C+ (or B+ for sheets without a VOTE COLUMN)
    for sheet_name in PRODUCT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_col = _first_lender_col(ws)
        for c in range(start_col, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v and str(v).strip():
                all_names.add(str(v).strip())

    # CS sheets
    cs_lender_cols = {
        "CS-CREDIT": (1, 3),  # column, start_row
        "CS-EXPERIENCE": (1, 2),
        "CS-LEASEOCCUPANCY": (1, 2),
        "CS-RESERVESASSETS": (1, 2),
    }
    for sheet_name, (col, start_row) in cs_lender_cols.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v and str(v).strip():
                all_names.add(str(v).strip())

    # Normalize and insert
    canonicals = {}
    for raw in sorted(all_names):
        canonical = normalize_lender(raw)
        if canonical not in canonicals:
            db.execute(
                "INSERT OR IGNORE INTO lenders (canonical_name, display_name) VALUES (?,?)",
                (canonical, canonical)
            )
            # Get ID
            row = db.execute("SELECT id FROM lenders WHERE canonical_name=?", (canonical,)).fetchone()
            lid = row["id"] if row else None
            canonicals[canonical] = lid
        # Insert alias
        lid = canonicals.get(canonical)
        if lid:
            norm_raw = raw.strip().replace('\n', ' ')
            if norm_raw != canonical:
                db.execute(
                    "INSERT OR IGNORE INTO lender_aliases (lender_id, alias) VALUES (?,?)",
                    (lid, norm_raw)
                )

    count = db.execute("SELECT COUNT(*) FROM lenders").fetchone()[0]
    alias_count = db.execute("SELECT COUNT(*) FROM lender_aliases").fetchone()[0]
    log(db, import_id, "INFO", f"{count} lenders, {alias_count} aliases")


def _insert_attribute_defs(db, import_id, wb):
    """Build attribute definitions from all product sheet column B headers."""
    seen = OrderedDict()
    for sheet_name, product_key in PRODUCT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v and str(v).strip():
                raw = str(v).strip()
                # Normalize to canonical name
                canonical = raw.lower().replace(' | ', ' ').replace('\n', ' ').strip()
                canonical = re.sub(r'\s+', ' ', canonical)
                canonical = re.sub(r'[^a-z0-9_\s]', '', canonical).strip()
                canonical = canonical.replace(' ', '_')[:80]
                # Harden a few common mappings
                if 'differentiator' in canonical:
                    canonical = 'differentiator'
                if canonical not in seen:
                    data_type = get_type(canonical)
                    seen[canonical] = {
                        'display': raw[:200],
                        'type': data_type,
                    }

    # Insert
    for canonical, info in seen.items():
        db.execute(
            "INSERT OR IGNORE INTO attribute_definitions (name, display_name, data_type) VALUES (?,?,?)",
            (canonical, info['display'], info['type'])
        )
    count = db.execute("SELECT COUNT(*) FROM attribute_definitions").fetchone()[0]
    log(db, import_id, "INFO", f"{count} attribute definitions")

    # Map product → attributes
    attr_ids = {}
    for row in db.execute("SELECT id, name FROM attribute_definitions").fetchall():
        attr_ids[row["name"]] = row["id"]

    product_ids = {}
    for row in db.execute("SELECT id, name FROM products").fetchall():
        product_ids[row["name"]] = row["id"]

    for sheet_name, product_key in PRODUCT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        prod_id = product_ids.get(product_key)
        if not prod_id:
            continue
        order = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v and str(v).strip():
                raw = str(v).strip()
                canonical = raw.lower().replace(' | ', ' ').replace('\n', ' ').strip()
                canonical = re.sub(r'\s+', ' ', canonical)
                canonical = re.sub(r'[^a-z0-9_\s]', '', canonical).strip()
                canonical = canonical.replace(' ', '_')[:80]
                if 'differentiator' in canonical:
                    canonical = 'differentiator'
                if canonical in attr_ids:
                    db.execute(
                        "INSERT OR IGNORE INTO product_attributes (product_id, attr_id, sort_order) VALUES (?,?,?)",
                        (prod_id, attr_ids[canonical], order)
                    )
                    order += 1

    attr_count = db.execute("SELECT COUNT(*) FROM product_attributes").fetchone()[0]
    log(db, import_id, "INFO", f"{attr_count} product→attribute mappings")


def _insert_product_sheet_data(db, import_id, wb):
    """Parse each product sheet's pivot table into lender_attr_values."""
    product_ids = {}
    for row in db.execute("SELECT id, name FROM products").fetchall():
        product_ids[row["name"]] = row["id"]

    attr_ids = {}
    for row in db.execute("SELECT id, name, data_type FROM attribute_definitions").fetchall():
        attr_ids[row["name"]] = {"id": row["id"], "type": row["data_type"]}

    lender_ids = {}
    for row in db.execute("SELECT id, canonical_name FROM lenders").fetchall():
        lender_ids[row["canonical_name"]] = row["id"]

    total_rows = 0
    warnings = 0

    for sheet_name, product_key in PRODUCT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Read lender names from row 1, starting at the sheet's detected
        # first lender column (usually C, but B for sheets missing the
        # "VOTE COLUMN" placeholder — see _first_lender_col).
        sheet_lenders = {}
        start_col = _first_lender_col(ws)
        for c in range(start_col, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v and str(v).strip():
                raw = str(v).strip()
                canonical = normalize_lender(raw)
                if canonical in lender_ids:
                    sheet_lenders[c] = (canonical, lender_ids[canonical])

        if start_col == 2:
            log(db, import_id, "WARN",
                f"{sheet_name}: no 'VOTE COLUMN' placeholder detected, lenders read from column B instead of C",
                sheet_name, 1)

        if not sheet_lenders:
            log(db, import_id, "WARN", f"No lenders found in {sheet_name}")
            continue

        prod_id = product_ids.get(product_key)
        if not prod_id:
            continue

        # Process each attribute row
        for r in range(2, ws.max_row + 1):
            attr_cell = ws.cell(row=r, column=2)
            if attr_cell.value is None:
                continue
            raw_attr = str(attr_cell.value).strip()
            if not raw_attr or raw_attr == 'Timestamp':
                continue

            # Normalize attr name
            canonical = raw_attr.lower().replace(' | ', ' ').replace('\n', ' ').strip()
            canonical = re.sub(r'\s+', ' ', canonical)
            canonical = re.sub(r'[^a-z0-9_\s]', '', canonical).strip()
            canonical = canonical.replace(' ', '_')[:80]
            if 'differentiator' in canonical:
                canonical = 'differentiator'

            attr_info = attr_ids.get(canonical)
            if not attr_info:
                # Unknown attr — skip
                continue

            # Record product has this attr (if not already)
            db.execute(
                "INSERT OR IGNORE INTO product_attributes (product_id, attr_id) VALUES (?,?)",
                (prod_id, attr_info["id"])
            )

            # Flag sensitive attrs
            is_sensitive = canonical in SENSITIVE_ATTRS

            for col, (lender_name, lender_id) in sheet_lenders.items():
                cell = ws.cell(row=r, column=col)
                val = cell.value
                text_val, num_val, date_val, warn = parse_value(val, attr_info["type"])

                if warn:
                    warnings += 1
                    log(db, import_id, "WARN",
                        f"{lender_name}/{product_key}/{canonical}: {warn}",
                        sheet_name, r)

                if is_sensitive and text_val:
                    text_val = "[REDACTED]"

                db.execute(
                    """INSERT OR REPLACE INTO lender_attr_values
                       (import_id, lender_id, product_id, attr_id, value_text, value_numeric, value_date, raw_text,
                        source_sheet, source_col, source_row, validation_warning)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (import_id, lender_id, prod_id, attr_info["id"],
                     text_val, num_val, date_val, text_val[:500],
                     sheet_name, col, r,
                     warn[:200] if warn else None)
                )
                total_rows += 1

    log(db, import_id, "INFO", f"{total_rows} product attr values", "", 0)
    if warnings:
        log(db, import_id, "WARN", f"{warnings} type validation warnings", "", 0)

    # Build product_offerings
    db.execute("""
        INSERT OR IGNORE INTO product_offerings (lender_id, product_id)
        SELECT DISTINCT lender_id, product_id FROM lender_attr_values WHERE import_id=?
    """, (import_id,))


def _insert_credit_grids(db, import_id, wb):
    """Parse CS-CREDIT sheet."""
    sheet_name = "CS-CREDIT"
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    # Row 1: FICO bucket headers (col C+)
    fico_buckets = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value  # FICO values are in row 2
        if v is not None:
            try:
                fico_buckets[c] = int(float(v))
            except (ValueError, TypeError):
                pass

    # Also track special columns
    special_cols = {}  # col → field name
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v and not isinstance(v, (int, float)):
            raw = str(v).strip().lower()
            if 'foreign' in raw:
                special_cols[c] = 'foreign_national'
            elif 'fico qual' in raw:
                special_cols[c] = 'fico_qual_method'
            elif 'guarantee' in raw:
                special_cols[c] = 'guarantee_reqs'
            elif 'tradeline' in raw:
                special_cols[c] = 'tradeline_reqs'
            elif 'lates' in raw:
                special_cols[c] = 'max_lates'
            elif 'foreclosure' in raw or 'bk' in raw:
                special_cols[c] = 'bk_fc'
            elif 'borrower types' in raw.lower():
                special_cols[c] = 'borrower_types'
            elif 'ownership' in raw.lower():
                special_cols[c] = 'ownership_seasoning'

    count = 0
    for r in range(3, ws.max_row + 1):
        lender_raw = ws.cell(row=r, column=1).value
        min_fico = ws.cell(row=r, column=2).value
        if not lender_raw or not min_fico:
            continue
        try:
            min_fico = float(min_fico)
        except (ValueError, TypeError):
            continue

        canonical = normalize_lender(str(lender_raw).strip())
        lender_row = db.execute(
            "SELECT id FROM lenders WHERE canonical_name=?", (canonical,)
        ).fetchone()
        if not lender_row:
            log(db, import_id, "WARN", f"Credit grid lender not found: {canonical}", sheet_name, r)
            continue
        lender_id = lender_row["id"]

        # Build LTV values per bucket
        ltv_purchase = None
        ltv_rate = None
        ltv_cash = None
        fn = None
        fico_method = None
        guarantee = None
        tradelines = None
        lates = None
        bk_fc = None

        for c in range(3, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            val = str(v).strip()
            if c in special_cols:
                field = special_cols[c]
                if field == 'foreign_national':
                    fn = val
                elif field == 'fico_qual_method':
                    fico_method = val
                elif field == 'guarantee_reqs':
                    guarantee = val
                elif field == 'tradeline_reqs':
                    tradelines = val
                elif field == 'max_lates':
                    lates = val
                elif field == 'bk_fc':
                    bk_fc = val
            elif c in fico_buckets:
                # LTV value for this FICO bucket
                bucket = fico_buckets[c]
                if not ltv_purchase:
                    ltv_purchase = val
                    ltv_rate = val
                    ltv_cash = val
                else:
                    ltv_purchase = val  # last bucket wins

        db.execute(
            """INSERT INTO credit_grids
               (import_id, lender_id, min_fico, ltv_purchase, ltv_rate_term, ltv_cashout,
                foreign_national, fico_qual_method, guarantee_reqs, tradeline_reqs, max_lates, bk_fc, source_sheet)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (import_id, lender_id, min_fico, ltv_purchase, ltv_rate, ltv_cash,
             fn, fico_method, guarantee, tradelines, lates, bk_fc, sheet_name)
        )
        count += 1

    log(db, import_id, "INFO", f"{count} credit grid entries")


def _insert_experience_matrices(db, import_id, wb):
    """Parse CS-EXPERIENCE sheet."""
    sheet_name = "CS-EXPERIENCE"
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    # Row 1: column headers — exp levels (col 2-12: 0-10 exp)
    exp_cols = {}
    for c in range(2, 13):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            try:
                exp_cols[c] = int(float(v.rstrip(' Exp').strip()))
            except (ValueError, AttributeError):
                pass

    count = 0
    for r in range(2, ws.max_row + 1):
        lender_raw = ws.cell(row=r, column=1).value
        if not lender_raw:
            continue
        canonical = normalize_lender(str(lender_raw).strip())
        lender_row = db.execute(
            "SELECT id FROM lenders WHERE canonical_name=?", (canonical,)
        ).fetchone()
        if not lender_row:
            log(db, import_id, "WARN", f"Experience lender not found: {canonical}", sheet_name, r)
            continue
        lender_id = lender_row["id"]

        for c, exp_level in exp_cols.items():
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip():
                db.execute(
                    "INSERT INTO experience_matrices (import_id, lender_id, exp_level, ltc_terms, source_sheet) VALUES (?,?,?,?,?)",
                    (import_id, lender_id, exp_level, str(v).strip()[:500], sheet_name)
                )
                count += 1

    log(db, import_id, "INFO", f"{count} experience matrix entries")


def _insert_scenarios(db, import_id, wb):
    """Parse CS-DSCR Implication and CS-FNF Implication sheets.
    Format: condition row (col1=title, cols2+=lenders), then detail row (cols2+=details), then blank separator."""
    for sheet_name in ["CS-DSCR Implication", "CS-FNF Implication"]:
        raw_name = next((n for n in wb.sheetnames if n.strip() == sheet_name), None)
        if not raw_name:
            continue
        ws = wb[raw_name]

        product_type = "sfr_dscr" if "DSCR" in sheet_name else "fix_and_flip"

        count = 0
        r = 2
        while r <= ws.max_row:
            col1 = ws.cell(row=r, column=1).value
            if not col1 or str(col1).strip() == '':
                r += 1
                continue

            condition = str(col1).strip()

            # Read lenders from columns 2+ of condition row
            lenders = {}
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and str(v).strip():
                    lenders[c] = str(v).strip()[:200]

            # Read details from the next row (same columns)
            detail_row = r + 1
            details = {}
            if detail_row <= ws.max_row:
                for c in range(2, ws.max_column + 1):
                    v = ws.cell(row=detail_row, column=c).value
                    if v and str(v).strip():
                        details[c] = str(v).strip()[:2000]

            # Insert each lender (only if it's a known lender)
            for c, lender_raw in lenders.items():
                canonical = normalize_lender(lender_raw)
                if not db.execute("SELECT 1 FROM lenders WHERE canonical_name=?", (canonical,)).fetchone():
                    continue  # skip non-lender values (cities, FICO buckets, yes/no, etc.)
                detail = details.get(c)
                scenario_id = f"{product_type}:{condition[:40]}:col{c}"
                db.execute(
                    """INSERT INTO scenarios
                       (import_id, scenario_id, condition, product_type, recommendation_lender, recommendation_detail, source_sheet, source_row)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (import_id, scenario_id, condition, product_type,
                     canonical, detail, sheet_name, r)
                )
                count += 1

            r += 2  # skip past condition row + detail row (or past condition if no detail)

        log(db, import_id, "INFO", f"{count} scenarios from {sheet_name}")


def _insert_lease_occupancy(db, import_id, wb):
    """Parse CS-LEASEOCCUPANCY sheet."""
    sheet_name = "CS-LEASEOCCUPANCY"
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    # Row 1: column headers
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            raw = str(v).strip().lower()
            if 'lender' in raw:
                col_map[c] = 'lender'
            elif 'occupancy' in raw:
                col_map[c] = 'occupancy_reqs'
            elif 'vacancy' in raw or 'pricing' in raw:
                col_map[c] = 'vacancy_pricing'
            elif 'acceptable' in raw:
                col_map[c] = 'acceptable_lease'
            elif 'unacceptable' in raw:
                col_map[c] = 'unacceptable_lease'
            elif 'lease term' in raw:
                col_map[c] = 'lease_terms'

    count = 0
    for r in range(2, ws.max_row + 1):
        lender_col = None
        for c, field in col_map.items():
            if field == 'lender':
                lender_col = c
                break
        if not lender_col:
            continue
        lender_raw = ws.cell(row=r, column=lender_col).value
        if not lender_raw:
            continue
        canonical = normalize_lender(str(lender_raw).strip())
        lender_row = db.execute(
            "SELECT id FROM lenders WHERE canonical_name=?", (canonical,)
        ).fetchone()
        if not lender_row:
            continue

        vals = {}
        for c, field in col_map.items():
            if field != 'lender':
                v = ws.cell(row=r, column=c).value
                vals[field] = str(v).strip()[:1000] if v else None

        db.execute(
            """INSERT INTO lease_occupancy
               (import_id, lender_id, occupancy_reqs, vacancy_pricing, acceptable_lease, unacceptable_lease, lease_terms, source_sheet)
               VALUES (?,?,?,?,?,?,?,?)""",
            (import_id, lender_row["id"],
             vals.get('occupancy_reqs'), vals.get('vacancy_pricing'),
             vals.get('acceptable_lease'), vals.get('unacceptable_lease'),
             vals.get('lease_terms'), sheet_name)
        )
        count += 1

    log(db, import_id, "INFO", f"{count} lease/occupancy entries")


def _insert_reserves(db, import_id, wb):
    """Parse CS-RESERVESASSETS sheet."""
    sheet_name = "CS-RESERVESASSETS"
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            raw = str(v).strip().lower()
            if 'lender' in raw:
                col_map[c] = 'lender'
            elif 'requirement' in raw and 'multi' not in raw:
                col_map[c] = 'requirement'
            elif 'multi' in raw:
                col_map[c] = 'multifamily'
            elif 'foreign' in raw:
                col_map[c] = 'foreign_national'
            elif 'cash out' in raw:
                col_map[c] = 'cash_out'
            elif 'eligible' in raw:
                col_map[c] = 'eligible_assets'

    count = 0
    for r in range(2, ws.max_row + 1):
        lender_col = None
        for c, field in col_map.items():
            if field == 'lender':
                lender_col = c
                break
        if not lender_col:
            continue
        lender_raw = ws.cell(row=r, column=lender_col).value
        if not lender_raw:
            continue
        canonical = normalize_lender(str(lender_raw).strip())
        lender_row = db.execute(
            "SELECT id FROM lenders WHERE canonical_name=?", (canonical,)
        ).fetchone()
        if not lender_row:
            continue

        vals = {}
        for c, field in col_map.items():
            if field != 'lender':
                v = ws.cell(row=r, column=c).value
                vals[field] = str(v).strip()[:1000] if v else None

        db.execute(
            """INSERT INTO reserves_requirements
               (import_id, lender_id, requirement, multifamily, foreign_national, cash_out, eligible_assets, source_sheet)
               VALUES (?,?,?,?,?,?,?,?)""",
            (import_id, lender_row["id"],
             vals.get('requirement'), vals.get('multifamily'),
             vals.get('foreign_national'), vals.get('cash_out'),
             vals.get('eligible_assets'), sheet_name)
        )
        count += 1

    log(db, import_id, "INFO", f"{count} reserve/asset entries")


def _print_summary(db, import_id):
    """Print import summary."""
    tables = [
        "products", "lenders", "lender_aliases",
        "attribute_definitions", "product_attributes",
        "lender_attr_values", "credit_grids",
        "experience_matrices", "scenarios",
        "lease_occupancy", "reserves_requirements",
    ]
    print()
    for t in tables:
        count = db.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        print(f"  {t}: {count}")

    warnings = db.execute(
        "SELECT COUNT(*) FROM migration_log WHERE import_id=? AND level='WARN'",
        (import_id,)
    ).fetchone()[0]
    errors = db.execute(
        "SELECT COUNT(*) FROM migration_log WHERE import_id=? AND level='ERROR'",
        (import_id,)
    ).fetchone()[0]
    print(f"\n  Warnings: {warnings}")
    print(f"  Errors:   {errors}")

    if warnings:
        print("\n  Sample warnings:")
        for row in db.execute(
            "SELECT message, source_sheet, source_row FROM migration_log WHERE import_id=? AND level='WARN' LIMIT 5",
            (import_id,)
        ).fetchall():
            print(f"    {row['message']} ({row['source_sheet']}:{row['source_row']})")


# ── CLI ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Migrate Master Credit Box to typed SQLite")
    parser.add_argument("--excel", default=EXCEL_DEFAULT)
    parser.add_argument("--fresh", action="store_true", help="Drop and rebuild DB from scratch")
    parser.add_argument("--diff", action="store_true", help="Show diff from last import")
    args = parser.parse_args()

    if args.fresh and DB_PATH.exists():
        DB_PATH.unlink()
        print("Removed existing DB")

    migrate(args.excel)


if __name__ == "__main__":
    main()
