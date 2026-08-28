"""
Parse Master Credit Box Excel → normalized JSON corpus.

Outputs:
  corpus/corpus.json        — all EAV triples + scenarios + profiles
  corpus/lenders.json       — lender index (names, aliases, contacts)
  corpus/scenarios.json     — decision rules (CS sheets)
"""

import json, re, os, sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl

SRC = Path(os.environ.get("CREDIT_BOX_SRC",
    "~/Downloads/Copy of THE Master Credit Box-IPLE 2026 (1).xlsx")).expanduser()
OUT = Path(__file__).parent / "corpus"

PRODUCT_SHEETS = {
    "SFR DSCR":            "sfr_dscr",
    "Fix & Flip":          "fix_and_flip",
    "New Construction":    "new_construction",
    "Multifamily Long Term":  "multifamily_lt",
    "SFR Bridge":          "sfr_bridge",
    "SFR Blanket":         "sfr_blanket",
    "Multifamily Rehab":   "multifamily_rehab",
    "Multi-Comm Bridge":   "multi_comm_bridge",
    "SB Commercial Long Term": "sb_commercial_lt",
}

CS_SHEETS = {
    "CS-DSCR Implication":    "decision_dscr",
    "CS-FNF Implication ":    "decision_fnf",
    "CS-CREDIT":              "credit_grid",
    "CS-EXPERIENCE":          "experience_matrix",
    "CS-RESERVESASSETS":      "reserves_policy",
    "CS-LEASEOCCUPANCY":      "lease_occupancy",
}

CANONICAL_ATTRS = {
    "fixed period": "fixed_period",
    "amortization period": "amortization",
    "interest only available": "io_available",
    "min loan amount": "loan_min",
    "max loan amount": "loan_max",
    "max % ltv (purchase)": "ltv_purchase_max",
    "max % ltv (rate/term refi)": "ltv_rateterm_max",
    "max % ltv (cash out refi)": "ltv_cashout_max",
    "fico requirement at max ltv": "fico_at_max_ltv",
    "min fico": "fico_min",
    "minimum fico": "fico_min",
    "fico min": "fico_min",
    "fico qualification": "fico_qualification",
    "floor rate ****": "rate_floor",
    "rate lock (when, how long)": "rate_lock",
    "lender origination": "origination_fee",
    "underwriting fee": "uw_fee",
    "processing fee": "processing_fee",
    "other lender fees: (ex 3rd party)": "other_fees",
    "prepayment penalty": "prepay_penalty",
    "broker / wholesale compensation": "broker_comp",
    "ysp buy up calc": "ysp_buyup",
    "ysp allowed to get to floor rate": "ysp_to_floor",
    "max % of purchase": "ltv_purchase_max",
    "max % ltv (purch.)": "ltv_purchase_max",
    "term in months": "term_months",
    "state coverage:": "state_coverage",
    "avg days to close": "days_to_close",
    "differentiator": "differentiator",
    "product": "product_name",
    "portfolio (y/n)": "is_portfolio",
    "max cash back while still r/t": "max_cashback_rateterm",
    "uses loanbidz credit pull?": "uses_loanbidz",
    "other tradeline requirements": "tradeline_reqs",
    "bk/fc/ mod in x years": "bk_fc_wait_years",
    "max mortgage lates": "max_lates",
    "dscr / prop dti (min /max)": "dscr_range",
    "asset requirement (mos pitia)": "reserves_pitia_months",
    "eligible asset types ": "eligible_assets",
    "purchase seasoning months to arv": "purchase_seasoning_months",
    "payment reserve (mos required) - collected at closing": "payment_reserve_months",
    "cash-out used for reserves": "cashout_as_reserves",
    "asset aging months": "asset_aging_months",
    "max exposure to individual guarantor": "max_guarantor_exposure",
    "allows seller 2nd liens": "seller_second_allowed",
    "wholesaler assignment fees": "wholesaler_assignment",
    "max seller contribution to closing costs": "seller_contribution_max",
    "entity types": "entity_types",
    "property types": "property_types",
    "occupancy requirements": "occupancy_reqs",
    "min property value": "property_value_min",
    "max property value": "property_value_max",
    "min lot size": "lot_size_min",
    "max acreage": "acreage_max",
    "max % ltv (ltc)": "ltc_max",
    "max % ltc": "ltc_max",
    "max ltv/ltc": "ltv_ltc_max",
    "min dscr": "dscr_min",
    "max dscr": "dscr_max",
    "interest rate": "rate_note",
    "minimum loan amount": "loan_min",
    "maximum loan amount": "loan_max",
    "max ltv": "ltv_max",
    "max ltc": "ltc_max",
    "ltv at max fico": "ltv_at_max_fico",
    "ltv at min fico": "ltv_at_min_fico",
    "rehab budget": "rehab_budget_note",
    "rate/term": "ltv_rateterm_note",
    "cash out": "ltv_cashout_note",
    "bridge/transition": "bridge_note",
    "ground up": "ground_up_note",
}

SENSITIVE_LABELS = {"user name", "password", "user", "pass", "login", "portal_url", "pricer_location",
                    "dropbox folder (pdf)", "dropbox folder (excel)", "lender dropbox folder link",
                    "other links", "website link"}


def cell_text(v):
    if v is None:
        return None
    return str(v).strip()


def is_numeric(v):
    if v is None:
        return False
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def normalize_label(s):
    if not s:
        return None
    s = re.sub(r'\*\*\*\*', '', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return CANONICAL_ATTRS.get(s, s.replace(" ", "_").replace("/", "_").replace("-", "_")[:50])


def get_lender_name(v):
    if not v or not isinstance(v, str):
        return None
    v = v.strip().rstrip('.')
    if v.upper() in ("VOTE COLUMN", "VOTE", "") or len(v) < 2:
        return None
    return v


def parse_product_sheet(ws, product_key):
    records = []
    max_row = ws.max_row
    max_col = ws.max_column

    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    # Detect layout: does col A hold labels or is it VOTE COLUMN?
    is_vote_sheet = str(row1[0] or "").strip() == "VOTE COLUMN"

    # Find label column and lender columns
    if is_vote_sheet:
        label_col = 1  # col B
        lenders = {}
        for c, v in enumerate(row1):
            name = get_lender_name(v)
            if name and c >= 2:
                lenders[c] = name
        # Row 2 column B might have the row label header
        label_header = str(row2[1] or "").strip() if len(row2) > 1 else ""
    else:
        label_col = 0  # col A
        lenders = {}
        for c, v in enumerate(row1):
            name = get_lender_name(v)
            if name and c >= 1:
                lenders[c] = name
        label_header = str(row1[0] or "").strip() if row1[0] else ""

    if not lenders:
        return records

    label_header_row = 2 if is_vote_sheet else 1

    start_row = 3 if is_vote_sheet else 2

    for r in range(start_row, max_row + 1):
        row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]

        # Get label from the label column
        raw_label = cell_text(row_vals[label_col]) if len(row_vals) > label_col else None
        if not raw_label:
            continue

        # Skip numeric-only labels (vote markers)
        if is_numeric(row_vals[label_col]):
            continue

        # Skip short markers
        if raw_label.lower() in ("updated",):
            continue

        attr_name = normalize_label(raw_label)
        is_sensitive = raw_label.lower().strip() in SENSITIVE_LABELS

        for c in lenders:
            if c >= len(row_vals):
                continue
            val = row_vals[c]
            if val is None:
                continue

            confidence = "typed" if (isinstance(val, (int, float)) and not isinstance(val, bool)) else "text"

            records.append({
                "lender": lenders[c],
                "product": product_key,
                "attr_name": attr_name,
                "attr_value": val if isinstance(val, (int, float)) else str(val).strip(),
                "raw_text": str(val).strip(),
                "source_sheet": ws.title,
                "source_row": r,
                "confidence": confidence,
                "sensitive": is_sensitive,
            })

    return records


def parse_cs_implication(ws, matrix_type):
    """Parse CS-DSCR Implication or CS-FNF Implication sheets."""
    scenarios = []
    max_row = ws.max_row
    max_col = ws.max_column

    r = 2
    while r <= max_row:
        row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        condition = cell_text(row_vals[0]) if row_vals else None
        if not condition or condition.lower() in ("updated", "standard deal"):
            r += 1
            continue

        # Next row has detail/notes
        if r + 1 <= max_row:
            detail_row = list(ws.iter_rows(min_row=r+1, max_row=r+1, values_only=True))[0]
        else:
            detail_row = []

        recommendations = []
        for c in range(1, max_col):
            lender = cell_text(row_vals[c]) if c < len(row_vals) and row_vals[c] else None
            detail = cell_text(detail_row[c]) if c < len(detail_row) and detail_row[c] else None
            if lender:
                recommendations.append({"lender": lender, "detail": detail or ""})

        if recommendations:
            slug = re.sub(r'[^a-z0-9]+', '_', condition.lower()).strip('_')
            scenarios.append({
                "scenario_id": f"{slug}",
                "condition": condition,
                "product_type": matrix_type,
                "recommendations": recommendations,
                "source_sheet": ws.title,
            })

        r += 2  # skip detail row
        # Skip blank separator rows
        while r <= max_row:
            nv = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if nv[0] is not None or any(v is not None for v in nv[1:4]):
                break
            r += 1

    return scenarios


def parse_cs_credit(ws):
    """Parse CS-CREDIT sheet — FICO grids per lender."""
    grids = []
    max_col = ws.max_column

    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    fico_buckets = {}
    for c in range(2, max_col):
        fico = row2[c]
        if is_numeric(fico):
            fico_buckets[c] = int(fico)
        elif isinstance(fico, str) and fico.strip().lower() == "minimum fico":
            fico_buckets[c] = "min"

    for r in range(3, ws.max_row + 1):
        row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        lender = cell_text(row_vals[0]) if row_vals else None
        if not lender:
            continue

        min_fico = row_vals[1] if len(row_vals) > 1 else None

        grid = {"lender": lender, "min_fico": min_fico, "source_sheet": ws.title}
        buckets = {}
        for c, bucket in fico_buckets.items():
            if c < len(row_vals):
                v = row_vals[c]
                if v is not None:
                    bucket_key = bucket
                    buckets[str(bucket_key)] = str(v).strip() if isinstance(v, str) else v
        if buckets:
            grid["grid"] = buckets
        else:
            grid["grid"] = {}

        grids.append(grid)

    return grids


def parse_cs_sheet(ws, cs_type):
    """Generic CS sheet parser — returns list of dicts."""
    rows = []
    headers = []
    header_row = None

    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        text_vals = [cell_text(v) for v in row_vals if v is not None]
        if len(text_vals) >= 3:
            header_row = r
            headers = [cell_text(v) for v in row_vals]
            break

    if header_row is None:
        return rows

    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        lender = cell_text(row_vals[0]) if row_vals else None
        if not lender:
            continue
        entry = {"lender": lender, "source_sheet": ws.title, "type": cs_type}
        for c, h in enumerate(headers):
            if c < len(row_vals) and row_vals[c] is not None:
                key = normalize_label(h) if h else f"col_{c}"
                entry[key] = str(row_vals[c]).strip() if isinstance(row_vals[c], str) else row_vals[c]
        rows.append(entry)

    return rows


def parse_sscs(ws):
    """Parse SSCS — lender aliases and contacts."""
    aliases = {}
    contacts = {}
    max_col = ws.max_column

    # Row 1 has lender names in odd columns (col 1, 3, 5, ...)
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    # Lenders start at column 1 (0-indexed, i.e. col B) and alternate
    for c in range(1, max_col, 2):
        name = get_lender_name(row1[c]) if c < len(row1) else None
        if not name:
            continue
        # Skip header labels like "Lenders"
        if name.lower() in ("lenders", "contacts"):
            continue

        aka = cell_text(row3[c]) if c < len(row3) and row3[c] else None
        if aka and aka.strip() and aka.strip().lower() not in ("aka's", "aka's "):
            aliased_name = aka.strip().rstrip('.')
            aliases[name] = {"canonical": name, "aka": [aliased_name]}
            aliases[aliased_name] = {"canonical": name, "aka": [aliased_name]}
        else:
            aliases[name] = {"canonical": name, "aka": []}

    # Row 2 has contacts in same odd columns
    for c in range(1, max_col, 2):
        name = get_lender_name(row1[c]) if c < len(row1) else None
        if not name or name.lower() in ("lenders", "contacts"):
            continue
        contact = cell_text(row2[c]) if c < len(row2) else None
        if contact:
            contacts[name] = contact

    return {"aliases": aliases, "contacts": contacts}


# Manual alias map from README analysis and cross-sheet comparison.
# (lender_variant → canonical_name)
MANUAL_ALIASES = {
    "CoreVest": "Corevest",
    "ShareStates": "Sharestates",
    "Lending One": "LendingOne",
    "ROC": "ROC Capital",
    "Visio Lending": "Visio",
    "Rodeo Capital": "Rodeo",
    "Rodeo": "Rodeo Capital",
    "RCN Capital": "RCN",
    "Kiavi Lending": "Kiavi",
    "Back Flip": "Backflip",
    "Backflip": "Back Flip",
    "Do Backflip": "Back Flip",
    "Aloha Capital": "Aloha",
    "Apex Lending A+": "Apex Lending",
    "Apex Lending": "Apex Lending",
    "IceCap Lending": "ICECap",
    "ICECap": "IceCap Lending",
    "Constructive (Standard)": "Constructive",
    "Constructive (Expanded)": "Constructive Expanded",
    "Constructive Draft": "Constructive",
    "Ground Floor": "Groundfloor",
    "Rain City": "Rain City Capital",
    "RCN (LoanBidz Special)": "RCN",
    "Easy Street": "EasyStreet",
    "Templeview": "Temple View Capital",
    "Wildcat": "Wildbird",
    "Crebird/Wildcat": "Wildbird",
    "Crebrid": "Wildbird",
    "Lima One": "Lima One Capital",
    "Flipco": "FlipCo Financial",
    "FlipCo": "FlipCo Financial",
    "Hard Money": "Hard Money Co",
    "Verus Capital": "Verus",
    "Corevest": "Corevest",
    "BPC (on pause)": "BPC",
    "Builder Finance": "Builders Capital",
}


def resolve_lender(name, alias_map):
    """Resolve alias to canonical name. Case-insensitive."""
    if not name:
        return name
    name = name.strip().rstrip('.')
    name_lower = name.lower()
    # Check manual aliases first (case-insensitive)
    for k, v in MANUAL_ALIASES.items():
        if k.lower() == name_lower:
            return v
    # Check auto-detected aliases from SSCS (case-insensitive)
    for k, v in alias_map.items():
        if k.lower() == name_lower:
            return v["canonical"]
    # Fuzzy: check if name starts with a known alias
    for k, v in MANUAL_ALIASES.items():
        if name_lower.startswith(k.lower().rstrip('.')) or k.lower().startswith(name_lower):
            return v
    return name


def main():
    print(f"Reading: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=True)

    OUT.mkdir(parents=True, exist_ok=True)

    all_records = []
    all_scenarios = []
    all_credit_grids = []
    all_cs_rows = []
    alias_data = None

    # 1. Product sheets
    for sheet_name, product_key in PRODUCT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP product sheet not found: {sheet_name}")
            continue
        ws = wb[sheet_name]
        records = parse_product_sheet(ws, product_key)
        all_records.extend(records)
        print(f"  {sheet_name}: {len(records)} attribute records")

    # 2. CS sheets
    for sheet_name, cs_type in CS_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP CS sheet not found: {sheet_name}")
            continue
        ws = wb[sheet_name]

        if cs_type in ("decision_dscr", "decision_fnf"):
            scenarios = parse_cs_implication(ws, cs_type)
            all_scenarios.extend(scenarios)
            print(f"  {sheet_name}: {len(scenarios)} scenarios")
        elif cs_type == "credit_grid":
            grids = parse_cs_credit(ws)
            all_credit_grids.extend(grids)
            print(f"  {sheet_name}: {len(grids)} lender grids")
        elif cs_type in ("lease_occupancy", "reserves_policy", "experience_matrix"):
            rows = parse_cs_sheet(ws, cs_type)
            all_cs_rows.extend(rows)
            print(f"  {sheet_name}: {len(rows)} lender entries")
        else:
            print(f"  UNKNOWN CS type: {sheet_name}")

    # 3. SSCS aliases
    if "SSCS" in wb.sheetnames:
        alias_data = parse_sscs(wb["SSCS"])
        print(f"  SSCS: {len(alias_data['aliases'])} aliases, {len(alias_data['contacts'])} contacts")

    # 4. Skip "Copy of CS-CREDIT"
    print("  Copy of CS-CREDIT: skipped (stale backup)")

    # Resolve lender aliases
    alias_map = alias_data["aliases"] if alias_data else {}
    for rec in all_records:
        resolved = resolve_lender(rec["lender"], alias_map)
        if resolved != rec["lender"]:
            rec["lender_canonical"] = resolved
        else:
            rec["lender_canonical"] = rec["lender"]

    for rec in all_credit_grids:
        rec["lender_canonical"] = resolve_lender(rec["lender"], alias_map)

    for rec in all_cs_rows:
        rec["lender_canonical"] = resolve_lender(rec["lender"], alias_map)

    for scenario in all_scenarios:
        for rec in scenario["recommendations"]:
            resolved = resolve_lender(rec["lender"], alias_map)
            if resolved != rec["lender"]:
                rec["lender_canonical"] = resolved
            else:
                rec["lender_canonical"] = rec["lender"]

    # Build lender index
    lender_names = set()
    for rec in all_records:
        lender_names.add(rec["lender_canonical"])
    for rec in all_credit_grids:
        lender_names.add(rec["lender_canonical"])
    for rec in all_cs_rows:
        lender_names.add(rec["lender_canonical"])
    for scenario in all_scenarios:
        for rec in scenario["recommendations"]:
            lender_names.add(rec["lender_canonical"])

    lenders_index = {}
    for name in sorted(lender_names):
        lenders_index[name] = {
            "name": name,
            "aliases": [k for k, v in alias_map.items() if v["canonical"] == name] if alias_map else [],
            "contact": alias_data["contacts"].get(name, "") if alias_data else "",
            "products": sorted(set(
                rec["product"] for rec in all_records
                if rec["lender_canonical"] == name
            )),
        }

    # Staleness tracking
    file_mtime = SRC.stat().st_mtime if SRC.exists() else 0
    file_age_days = (datetime.now().timestamp() - file_mtime) / 86400 if file_mtime else 0
    file_mtime_str = datetime.fromtimestamp(file_mtime).isoformat() if file_mtime else ""

    # Write outputs
    corpus = {
        "meta": {
            "generated": datetime.now().isoformat(),
            "source": str(SRC),
            "file_mtime": file_mtime_str,
            "file_age_days": round(file_age_days, 1),
            "file_size_bytes": SRC.stat().st_size if SRC.exists() else 0,
            "product_sheets": list(PRODUCT_SHEETS.keys()),
        },
        "records": all_records,
        "scenarios": all_scenarios,
        "credit_grids": all_credit_grids,
        "underwriting": all_cs_rows,
    }

    with open(OUT / "corpus.json", "w") as f:
        json.dump(corpus, f, indent=2, default=str)
    print(f"\nWrote: {OUT / 'corpus.json'} ({len(all_records)} records, {len(all_scenarios)} scenarios, {len(all_credit_grids)} credit grids, {len(all_cs_rows)} uw entries)")

    with open(OUT / "lenders.json", "w") as f:
        json.dump(lenders_index, f, indent=2)
    print(f"Wrote: {OUT / 'lenders.json'} ({len(lenders_index)} lenders)")

    scenarios_out = {
        "meta": {"generated": datetime.now().isoformat(), "source": str(SRC)},
        "scenarios": all_scenarios,
    }
    with open(OUT / "scenarios.json", "w") as f:
        json.dump(scenarios_out, f, indent=2, default=str)
    print(f"Wrote: {OUT / 'scenarios.json'} ({len(all_scenarios)} scenarios)")

    print("\nDone.")


if __name__ == "__main__":
    main()
